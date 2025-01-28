import argparse
from ..model.constant import Q_IDX
import os
from typing import Dict, Tuple
from h5dataloader.common.structure import *
from h5dataloader.common.common import parse_colors, parse_src_dst
import numpy as np
import h5py
from tqdm import tqdm
import cv2

from openpyxl.worksheet.worksheet import Worksheet
import openpyxl as xl

from ..model.constant import *
from ..model.metric_numpy import *
import json

import warnings
warnings.filterwarnings('ignore')

TAG_HEIGHT = 32
PADDING = 4
FONT_SCALE = 0.8
FONT_COLOR = (255, 255, 255)
FONT_FACE = cv2.FONT_HERSHEY_SIMPLEX
FONT_THICKNESS = 2

def parse_args() -> Dict[str, str]:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        '-i', '--input',
        type=str, metavar='PATH', required=True,
        help='Input path.'
    )
    parser.add_argument(
        '-o', '--output',
        type=str, metavar='PATH', default=None,
        help='Output path. Default is "[input dir]/data.avi"'
    )
    parser.add_argument(
        f'--{arg_hyphen(ARG_THRESHOLDS)}',
        type=float, nargs='+', default=DEFAULT_THRESHOLDS,
        help='Thresholds of depth.'
    )
    parser.add_argument(
        '-c', '--config',
        type=str, metavar='PATH', default=None,
        help='evaluation config file path.'
    )
    args = vars(parser.parse_args())

    if isinstance(args['output'], str):
        if os.path.isdir(os.path.dirname(args['output'])) is False:
            raise NotADirectoryError(os.path.dirname(args['output']))
    else:
        output_dir: str = os.path.dirname(args['input'])
        args['output'] = os.path.join(output_dir, 'reresult.xlsx')
    return args

def convert_label(src:np.ndarray, label_tag:str, label_convert_configs: Dict) -> np.ndarray:
    """convert_label

    ラベルの変換を行う

    Args:
        src (np.ndarray): 変換するNumpy行列
        label_tag (str): ラベルの設定のタグ

    Returns:
        np.ndarray: 変換後のNumpy行列
    """
    tmp:np.ndarray = src.copy()
    for label_convert_config in label_convert_configs[label_tag]:
        tmp = np.where(src == label_convert_config[CONFIG_TAG_SRC], label_convert_config[CONFIG_TAG_DST], tmp)
    return tmp

def _excel_detail_output(idx: int, metric_results: Dict[str, METRIC_RESULT], excel_sheet: Worksheet) -> int:
        itr: int = idx
        row: int = itr + 3
        col: int = 1
        excel_sheet.cell(row=row, column=col, value=itr)
        col += 1
        for metric_key, metric_result in metric_results.items():
            if metric_key == METRIC_IOU:
                value: float = metric_result.all_metric.item() * 100.0
            else:
                value: float = metric_result.all_metric.item()
            excel_sheet.cell(row=row, column=col, value=value)
            col += 1

            if metric_result.use_class is True:
                for j in range(metric_result.class_metric.shape[0]):
                    if metric_key == METRIC_IOU:
                        value: float = metric_result.class_metric[j].item(
                        ) * 100.0
                    else:
                        value: float = metric_result.class_metric[j].item()
                    excel_sheet.cell(row=row, column=col, value=value)
                    col += 1

def _excel_overview_output(metric_results: Dict[str, METRIC_RESULT], excel_sheet: Worksheet, depth_norm_range_max: float, result_dict: Dict[str, float] = None):
    for row, (metric_key, metric_result) in enumerate(metric_results.items(), 2):
        excel_sheet.cell(row=row, column=1, value=metric_result.name)

        if metric_key == METRIC_IOU:
            mean_all: np.ndarray = (
                metric_result.sum_class_metric / metric_result.counts_metric).mean()
            mean_all_value: float = mean_all.item() * 100.0
        else:
            mean_all: np.ndarray = metric_result.sum_class_metric.sum() / \
                metric_result.counts_metric.sum()

            if metric_key == METRIC_RMSE:
                mean_all_value: float = np.sqrt(
                    mean_all).item() * depth_norm_range_max
            else:
                mean_all_value: float = mean_all.item()

        excel_sheet.cell(row=row, column=2, value=mean_all_value)
        if isinstance(result_dict, dict):
            result_dict[metric_key] = mean_all_value

        if metric_result.use_class is True:
            col: int = 3

            if metric_result.use_class is True:
                if metric_result.counts_metric.ndim == 2:
                    mean_class: torch.Tensor = metric_result.sum_class_metric / \
                        metric_result.counts_metric.sum(axis=0)
                else:
                    mean_class: torch.Tensor = metric_result.sum_class_metric / \
                        metric_result.counts_metric

                for i in range(mean_class.shape[0]):
                    if metric_key == METRIC_RMSE:
                        excel_sheet.cell(row=row, column=col, value=np.sqrt(
                            mean_class[i]).item() * depth_norm_range_max)
                    elif metric_key == METRIC_IOU:
                        excel_sheet.cell(row=row, column=col,
                                        value=mean_class[i].item() * 100.0)
                    else:
                        excel_sheet.cell(row=row, column=col,
                                        value=mean_class[i].item())
                    col += 1

            if metric_result.use_thresholds is True:
                if metric_result.counts_metric.ndim == 2:
                    mean_threshold: np.ndarray = metric_result.sum_threshold_metric / \
                        metric_result.counts_metric.sum(axis=1)
                else:
                    mean_threshold: np.ndarray = metric_result.sum_threshold_metric / \
                        metric_result.counts_metric

                for i in range(mean_threshold.shape[0]):
                    if metric_key == METRIC_RMSE:
                        excel_sheet.cell(row=row, column=col, value=np.sqrt(
                            mean_threshold[i]).item() * depth_norm_range_max)
                    else:
                        excel_sheet.cell(row=row, column=col,
                                        value=mean_threshold[i].item())
                    col += 1

class Metric_ReResult:
    def __init__(self, args):
        self.input_hdf5 = args['input']
        self.config = args['config']
        self.output = args['output']
        
        self.label_convert_configs:Dict[str, List[Dict[str, int]]] = {}                     # ラベルの変換の設定を格納した辞書
        self.label_color_configs:Dict[str, List[Dict[str, Union[int, np.ndarray]]]] = {}    # ラベルの色の設定を格納した辞書

        self.num_classes: int = 0
        self.minibatch:Dict[str, Dict[str, Union[str, Dict[str, str], List[int], bool, List[float]]]] = {}     # mini-batchの設定

        self._load_config()
        self.main(args)

    def _load_config(self) -> None:
        # configファイルが見つからない場合に終了
        if os.path.isfile(self.config) is False:
            print('File not found : "%s"'%(self.config))
            exit(1)
            
        # configファイルをロード -> 辞書型へ
        config_dict:Dict[str, Dict[str, Dict[str, Any]]] = {}
        with open(self.config, mode='r') as configfile:
            config_dict = json.load(configfile)

        # ラベルの設定の辞書を生成
        if CONFIG_TAG_LABEL in config_dict:
            for key_label, item_label in config_dict[CONFIG_TAG_LABEL][CONFIG_TAG_CONFIG].items():
                # print('Label "%s"'%(key_label))
                self.label_convert_configs[key_label] = parse_src_dst(item_label, quiet=True)
                self.label_color_configs[key_label] = parse_colors(item_label)
                
        self.num_classes = len(self.label_color_configs[config_dict[CONFIG_TAG_MINIBATCH][DATASET_LABEL][CONFIG_TAG_LABELTAG]])

        for key, item in config_dict[CONFIG_TAG_MINIBATCH].items():
            data_dict:Dict[str, Dict[str, Any]] = {}
            data_dict[CONFIG_TAG_FROM] = item.get(CONFIG_TAG_FROM)
            data_dict[CONFIG_TAG_TYPE] = item.get(CONFIG_TAG_TYPE)

            if data_dict[CONFIG_TAG_FROM] is None or data_dict[CONFIG_TAG_TYPE] is None:
                print('keys "from" and "type" must not be null')
                exit(1)

            if isinstance(item.get(CONFIG_TAG_SHAPE), list) is True:
                data_dict[CONFIG_TAG_SHAPE] = tuple(item[CONFIG_TAG_SHAPE])
            else:
                data_dict[CONFIG_TAG_SHAPE] = None

            if isinstance(item.get(CONFIG_TAG_RANGE), list) is True:
                depth_range = []
                depth_range.append(0.0 if item[CONFIG_TAG_RANGE][0] < 0.0 else item[CONFIG_TAG_RANGE][0])
                depth_range.append(np.inf if item[CONFIG_TAG_RANGE][1] is None else item[CONFIG_TAG_RANGE][1])
                data_dict[CONFIG_TAG_RANGE] = tuple(depth_range)
            else:
                data_dict[CONFIG_TAG_RANGE] = DEFAULT_RANGE[data_dict[CONFIG_TAG_TYPE]]

            data_dict[CONFIG_TAG_NORMALIZE] = item.get(CONFIG_TAG_NORMALIZE)
            data_dict[CONFIG_TAG_LABELTAG] = item.get(CONFIG_TAG_LABELTAG)

            self.minibatch[key] = data_dict

    def main(self, args):
        ##################
        # Metric Setting #
        ##################
        label_color_configs: List[Dict[str, Union[str, int]]
                            ] = self.label_color_configs[self.minibatch[DATASET_LABEL][CONFIG_TAG_LABELTAG]]
        args[ARG_LABEL_TAGS] = {
            str(lcc[CONFIG_TAG_LABEL]): lcc[CONFIG_TAG_TAG] for lcc in label_color_configs}
        # set depth normalization max
        depth_norm_range_max = self.minibatch[DATASET_DEPTH][CONFIG_TAG_RANGE][1] \
            if self.minibatch[DATASET_DEPTH][CONFIG_TAG_NORMALIZE] is True else 1.0
        thresholds: List[float] = args[ARG_THRESHOLDS]
        normalized_thresholds: List[float] = [th / depth_norm_range_max for th in thresholds]
        
        metricIntersection = MetricIntersection()
        metricUnion = MetricUnion()
        metricBatchSumAE = MetricBatchSumAE(
            thresholds=normalized_thresholds, num_classes=self.num_classes)
        metricBatchSumSE = MetricBatchSumSE(
            thresholds=normalized_thresholds, num_classes=self.num_classes)
        metricBatchSumAPE = MetricBatchSumAPE(
            thresholds=normalized_thresholds, num_classes=self.num_classes)
        
        label_tag:str = self.minibatch[DATASET_LABEL][CONFIG_TAG_LABELTAG]

        if label_tag not in self.label_convert_configs.keys():
            print(f'"{label_tag}" is not in "/label"')
            exit(1)

        # set metric results
        metric_results: Dict[str, METRIC_RESULT_NUMPY] = {
            METRIC_IOU: METRIC_RESULT_NUMPY(name='IoU [%]', use_class=True, all_tag='Mean'),
            METRIC_MAE: METRIC_RESULT_NUMPY(name='MAE [m]', use_class=True, use_thresholds=True),
            METRIC_RMSE: METRIC_RESULT_NUMPY(name='RMSE [m]', use_class=True, use_thresholds=True),
            METRIC_MAPE: METRIC_RESULT_NUMPY(name='MAPE', use_class=True, use_thresholds=True),
        }

        # Excel Book
        eval_book = xl.Workbook()
        eval_sheet_overview = eval_book.worksheets[0]

        eval_sheet_overview.title = 'Overview'
        eval_sheet_overview.cell(row=1, column=1, value='Metric')
        eval_sheet_overview.cell(row=1, column=2, value='All')
        for col, (label, tag) in enumerate(args[ARG_LABEL_TAGS].items(), 3):
            eval_sheet_overview.cell(row=1, column=col, value=f'{label}:{tag}')
        tmp_thresholds = [0.0] + args[ARG_THRESHOLDS]
        for col, (th_min, th_max) in enumerate(zip(tmp_thresholds[:-1], tmp_thresholds[1:]), col+1):
            eval_sheet_overview.cell(
                row=1, column=col, value=f'{th_min}~{th_max}[m]')

        eval_sheet_detail = eval_book.create_sheet('Detail')
        eval_sheet_detail.cell(row=1, column=1, value='Step')
        eval_sheet_detail.merge_cells(
            start_row=1, end_row=2, start_column=1, end_column=1)

        metric_col = 2
        for metric_result in metric_results.values():
            eval_sheet_detail.cell(row=1, column=metric_col,
                                value=metric_result.name)
            if metric_result.use_class is True:
                cols: int = len(args[ARG_LABEL_TAGS]) + 1
                eval_sheet_detail.merge_cells(
                    start_row=1, end_row=1, start_column=metric_col, end_column=metric_col+cols-1)
                eval_sheet_detail.cell(
                    row=2, column=metric_col, value=metric_result.all_tag)
                for col, (label, tag) in enumerate(args[ARG_LABEL_TAGS].items(), metric_col + 1):
                    eval_sheet_detail.cell(
                        row=2, column=col, value=f'{label}:{tag}')
            else:
                cols: int = 1
                eval_sheet_detail.merge_cells(
                    start_row=1, end_row=2, start_column=metric_col, end_column=metric_col)
            metric_col += cols

        result_dict: Dict[str, Union[float, str]] = {DIR_RESULTS: os.path.dirname(args['output'])}

        try:
            with h5py.File(args['input'], mode='r') as h5file:
                data_group: h5py.Group = h5file['data']

                for itr in tqdm(range(h5file['header/length'][()]), desc='HDF5 -> ReResult'):
                    src_group: h5py.Group = data_group[str(itr)]

                    # in_camera: np.ndarray = src_group['Input-Camera'][()]
                    # in_map: np.ndarray = depth2colormap(
                        # src_group['Input-Map'][()], 0.0, 100.0)
                    # reevaluate using only convert label
                    pr_seg: np.ndarray = convert_label(src_group['Pred-Label'][()], label_tag, self.label_convert_configs)
                    gt_seg:np.ndarray = convert_label(src_group['GT-Label'][()], label_tag, self.label_convert_configs)
                    pr_depth: np.ndarray = src_group['Pred-Depth'][()] / depth_norm_range_max
                    gt_depth: np.ndarray = src_group['GT-Depth'][()] / depth_norm_range_max

                    # AE/MAE
                    ae: METRIC_SUM_OUTPUT = metricBatchSumAE(
                        pr_depth, gt_depth, gt_seg)
                    metric_results[METRIC_MAE].class_metric = ae.sum.sum(
                            axis=0) / ae.count.sum(axis=0) * depth_norm_range_max
                    metric_results[METRIC_MAE].threshold_metric = ae.sum.sum(
                        axis=1) / ae.count.sum(axis=1) * depth_norm_range_max
                    metric_results[METRIC_MAE].all_metric = ae.sum.sum(
                        axis=(0, 1)) / ae.count.sum(axis=(0, 1)) * depth_norm_range_max
                    metric_results[METRIC_MAE].add(ae.count, ae.sum.sum(
                        axis=0) * depth_norm_range_max, ae.sum.sum(axis=1) * depth_norm_range_max)
                    
                    # SE/RMSE
                    se: METRIC_SUM_OUTPUT = metricBatchSumSE(
                        pr_depth, gt_depth, gt_seg)
                    metric_results[METRIC_RMSE].class_metric = np.sqrt(
                        se.sum.sum(axis=0) / se.count.sum(axis=0)) * depth_norm_range_max
                    metric_results[METRIC_RMSE].threshold_metric = np.sqrt(
                        se.sum.sum(axis=1) / se.count.sum(axis=1)) * depth_norm_range_max
                    metric_results[METRIC_RMSE].all_metric = np.sqrt(
                        se.sum.sum(axis=(0, 1)) / se.count.sum(axis=(0, 1))) * depth_norm_range_max
                    metric_results[METRIC_RMSE].add(se.count, 
                        se.sum.sum(axis=0), se.sum.sum(axis=1))
                    
                    # APE/MAPE
                    ape: METRIC_SUM_OUTPUT = metricBatchSumAPE(
                        pr_depth, gt_depth, gt_seg)
                    metric_results[METRIC_MAPE].class_metric = ape.sum.sum(
                        axis=0) / ape.count.sum(axis=0)
                    metric_results[METRIC_MAPE].threshold_metric = ape.sum.sum(
                        axis=1) / ape.count.sum(axis=1)
                    metric_results[METRIC_MAPE].all_metric = ape.sum.sum(
                        axis=(0, 1)) / ape.count.sum(axis=(0, 1))
                    metric_results[METRIC_MAPE].add(ape.count, 
                        ape.sum.sum(axis=0), ape.sum.sum(axis=1))

                    # IoU
                    intersection: np.ndarray = metricIntersection(
                        pr_seg, gt_seg, self.num_classes)
                    union: np.ndarray = metricUnion(pr_seg, gt_seg, self.num_classes)
                    metric_results[METRIC_IOU].class_metric = intersection / union
                    metric_results[METRIC_IOU].all_metric = np.mean(metric_results[METRIC_IOU].class_metric, axis=0)
                    metric_results[METRIC_IOU].add(
                        union, intersection)

                    # -> Excel (Detail)
                    _excel_detail_output(
                        idx=itr, metric_results=metric_results, excel_sheet=eval_sheet_detail)

                # # -> Excel (Overview)
                _excel_overview_output(
                    metric_results, eval_sheet_overview, depth_norm_range_max, result_dict)
        except KeyboardInterrupt:
            print('KeyboardInterrupt')
        finally:
            eval_book.save(self.output)
            print(f'Saved: {args["output"]}')
            


if __name__ == '__main__':
    args = parse_args()
    reresult = Metric_ReResult(args)
